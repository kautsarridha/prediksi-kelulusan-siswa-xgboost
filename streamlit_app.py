import streamlit as st
import pandas as pd
import numpy as np
import joblib
import os
import io

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.colors import black, lightgrey, HexColor

from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font

# =====================
# PAGE CONFIG
# =====================
st.set_page_config("Prediksi Kelulusan Siswa", layout="wide")

# =====================
# UI STYLE + ANIMATION
# =====================
st.markdown("""
<style>

/* === HAPUS SEMUA PANEL / CARD KOSONG === */
div[data-testid="stVerticalBlock"]:empty {
    display: none !important;
    height: 0 !important;
    padding: 0 !important;
    margin: 0 !important;
}

.card:empty {
    display: none !important;
    height: 0 !important;
    padding: 0 !important;
    margin: 0 !important;
}

/* Hapus container kosong bawaan streamlit */
section.main > div:has(> div:empty) {
    display: none !important;
}

/* ======================= */

@keyframes fadeIn {
  from {opacity:0; transform: translateY(15px);}
  to {opacity:1; transform: translateY(0);}
}

@keyframes float {
  0% {transform: translateY(0px);}
  50% {transform: translateY(-6px);}
  100% {transform: translateY(0px);}
}

.stApp {
  background: radial-gradient(circle at top, #0A1828);
  color: #BFA181;
}

.main-title {
  font-size: 46px;
  font-weight: 800;
  animation: fadeIn 1s ease;
  text-align: center;
}

.subtitle {
  color:#9ca3af;
  margin-bottom:30px;
  animation: fadeIn 1.2s ease;
  text-align: center;
}

.card {
  background: rgba(15,23,42,0.75);
  backdrop-filter: blur(10px);
  border-radius: 18px;
  padding: 26px;
  margin-bottom: 28px;
  border: 1px solid rgba(255,255,255,0.08);
  box-shadow: 0 0 35px rgba(0,0,0,0.45);
  animation: fadeIn 0.9s ease;
}

.card:hover {
  transform: translateY(-4px) scale(1.01);
  transition: 0.3s ease;
  box-shadow: 0 0 55px rgba(37,99,235,0.35);
}

.stButton button, .stDownloadButton button {
  background: linear-gradient(135deg,#2563eb,#4f46e5);
  color: white;
  border-radius: 10px;
  font-weight: 600;
  padding: 10px 18px;
  transition: 0.25s;
}

.stButton button:hover, .stDownloadButton button:hover {
  transform: scale(1.05);
  box-shadow: 0 0 20px rgba(79,70,229,0.6);
}

.metric-card {
  background: linear-gradient(135deg,#1e293b,#020617);
  border-radius: 14px;
  padding: 18px;
  text-align:center;
  animation: float 4s ease-in-out infinite;
  border: 1px solid rgba(255,255,255,0.06);
}

.footer {
  text-align:center;
  color:#94a3b8;
  margin-top:60px;
  font-size:13px;
  opacity:0.8;
}
</style>
""", unsafe_allow_html=True)

# =====================
# HEADER
# =====================
st.markdown('<div class="main-title">🎓 Prediksi Kelulusan Siswa</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle">Untuk Memprediksi Tingkat Kelulusan Pada Siswa • Machine Learning XGBoost</div>', unsafe_allow_html=True)

# =====================
# TEMPLATE GENERATOR
# =====================
def generate_template_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "DATA"

    headers = ["NO","NAMA","SEMESTER 1","SEMESTER 2","SEMESTER 3","SEMESTER 4"]
    ws.append(headers)

    sample = [
        [1,"Nama Siswa 1",80,82,84,86],
        [2,"Nama Siswa 2",85,87,88,89],
        [3,"Nama Siswa 3",90,91,92,93],
    ]

    for row in sample:
        ws.append(row)

    thin = Side(border_style="thin")
    border = Border(left=thin,right=thin,top=thin,bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# =====================
# TEMPLATE CARD
# =====================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.subheader("📥 Download Template Excel")
st.write("Gunakan template ini agar format sesuai sistem prediksi.")

st.download_button(
    "📊 Download Template Excel Prediksi",
    data=generate_template_excel(),
    file_name="template_prediksi_kelulusan.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
st.markdown('</div>', unsafe_allow_html=True)

# =====================
# LOAD MODEL
# =====================
@st.cache_resource
def load_model():
    base = os.path.dirname(os.path.abspath(__file__))
    return joblib.load(os.path.join(base,"model_xgboost.pkl"))

try:
    model = load_model()
except:
    st.error("❌ model_xgboost.pkl tidak ditemukan")
    st.stop()

# =====================
# UPLOAD CARD
# =====================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.subheader("📂 Upload File Excel Nilai Siswa")
uploaded_file = st.file_uploader("Upload file Excel (.xlsx)", type=["xlsx"])
st.markdown('</div>', unsafe_allow_html=True)

if uploaded_file:
    try:
        df = pd.read_excel(uploaded_file)
    except:
        st.error("❌ Gagal membaca file Excel")
        st.stop()

    df.columns = [str(c).strip().upper() for c in df.columns]

    if df.columns.str.contains("UNNAMED").any():
        df = df.iloc[:, :6]
        df.columns = ["NO","NAMA","SEMESTER 1","SEMESTER 2","SEMESTER 3","SEMESTER 4"]

    semester_cols = ["SEMESTER 1","SEMESTER 2","SEMESTER 3","SEMESTER 4"]

    if not all(col in df.columns for col in semester_cols):
        st.error("❌ Format Excel tidak sesuai template")
        st.stop()

    for col in semester_cols:
        df[col] = df[col].astype(str).str.replace(",",".",regex=False)
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df.dropna(subset=semester_cols)

    X = df[semester_cols].values
    y_pred = model.predict(X)

    label_map = {0:"Berpotensi Tidak Lulus",1:"Lulus Baik",2:"Lulus Sangat Baik"}

    df["HASIL PREDIKSI"] = [label_map[i] for i in y_pred]
    df["RATA-RATA"] = df[semester_cols].mean(axis=1).round(2)

    st.markdown("""
<div class="card">
<h4>⚠️ Disclaimer Sistem Prediksi</h4>
<p style="color:#cbd5e1; line-height:1.6;">
Sistem ini menggunakan algoritma <b>Machine Learning (XGBoost)</b> untuk membantu memprediksi tingkat kelulusan siswa berdasarkan data nilai akademik.
<br><br>
Hasil prediksi <b>bukan keputusan final</b> dan <b>tidak menjamin akurasi 100%</b>.  
Keputusan kelulusan tetap harus dilakukan melalui evaluasi manual oleh pihak sekolah atau guru yang berwenang.
</p>
</div>
""", unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.subheader("📊 Hasil Prediksi")
    st.dataframe(df, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

    total = len(df)
    tdk = (df["HASIL PREDIKSI"]=="Berpotensi Tidak Lulus").sum()
    baik = (df["HASIL PREDIKSI"]=="Lulus Baik").sum()
    sb = (df["HASIL PREDIKSI"]=="Lulus Sangat Baik").sum()

    c1,c2,c3,c4 = st.columns(4)
    c1.markdown(f'<div class="metric-card">👥<br><b>{total}</b><br>Total</div>', unsafe_allow_html=True)
    c2.markdown(f'<div class="metric-card">⚠️<br><b>{tdk}</b><br>Berpotensi Tidak Lulus</div>', unsafe_allow_html=True)
    c3.markdown(f'<div class="metric-card">✅<br><b>{baik}</b><br>Lulus Baik</div>', unsafe_allow_html=True)
    c4.markdown(f'<div class="metric-card">🌟<br><b>{sb}</b><br>Sangat Baik</div>', unsafe_allow_html=True)

    if st.button("📄 Download PDF Laporan"):
        pdf_path = "Laporan_Prediksi_Kelulusan.pdf"
        doc = SimpleDocTemplate(pdf_path,pagesize=landscape(A4))

        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Laporan Prediksi Kelulusan Siswa (XGBoost)", styles["Title"]))
        elements.append(Paragraph(f"Total Siswa: {total}", styles["Normal"]))
        elements.append(Paragraph(f"Berpotensi Tidak Lulus: {tdk}", styles["Normal"]))
        elements.append(Paragraph(f"Lulus Baik: {baik}", styles["Normal"]))
        elements.append(Paragraph(f"Lulus Sangat Baik: {sb}", styles["Normal"]))
        header = ["NO","NAMA"]+semester_cols+["RATA-RATA","HASIL PREDIKSI"]
        table_data = [header]

        for _,row in df.iterrows():
            table_data.append([
                row["NO"],row["NAMA"],
                row["SEMESTER 1"],row["SEMESTER 2"],row["SEMESTER 3"],row["SEMESTER 4"],
                row["RATA-RATA"],row["HASIL PREDIKSI"]
            ])

        table = Table(table_data, repeatRows=1)

        style = TableStyle([
            ("BACKGROUND",(0,0),(-1,0),lightgrey),
            ("GRID",(0,0),(-1,-1),0.5,black),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("FONTSIZE",(0,0),(-1,-1),7),
        ])

        for i in range(1,len(table_data)):
            status = table_data[i][-1]
            if status=="Berpotensi Tidak Lulus":
                style.add("BACKGROUND",(-1,i),(-1,i),HexColor("#FFF3A0"))
            elif status=="Lulus Baik":
                style.add("BACKGROUND",(-1,i),(-1,i),HexColor("#CCE5FF"))
            else:
                style.add("BACKGROUND",(-1,i),(-1,i),HexColor("#C8F7C5"))

        table.setStyle(style)
        elements.append(table)
        doc.build(elements)

        with open(pdf_path,"rb") as f:
            st.download_button("⬇ Download PDF Sekarang",f,file_name=pdf_path,mime="application/pdf")

# =====================
# FOOTER
# =====================
st.markdown("""
<div class="footer">
Sistem Prediksi Kelulusan Siswa <br>
Streamlit • XGBoost • 2026
</div>
""", unsafe_allow_html=True)
